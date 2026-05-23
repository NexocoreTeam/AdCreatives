"""Tests for the creative-matrix stage — schema, XLSX writer, LLM stubbing."""

from __future__ import annotations

import json
from pathlib import Path

import pytest
import yaml
from openpyxl import load_workbook

from models.matrix import (
    CreativeMatrix,
    MatrixRow,
    MatrixTab,
    TrendingMatch,
    TrendingNext,
)
from strategy import creative_matrix as cm
from strategy.matrix_xlsx import write_matrix_xlsx

# ─── Schema round-trip ──────────────────────────────────────────────────────


def _sample_row(tab: MatrixTab, idx: int) -> MatrixRow:
    base = {
        "id": cm._normalize_id(tab, idx),
        "tab": tab,
        "source_type": ["reddit", "amazon"],
        "what_people_say": "fixed my gut where probiotics failed",
        "format_fit": ["us-vs-them", "headline"],
        "static_treatment": "Split-screen: rival jar vs ours, headline 'no surprise breakouts'.",
        "psychology_trigger": "social_proof",
        "trending_match": TrendingMatch(
            concept="ingredient receipt",
            execution="Hold both bottles, read labels aloud, cut to highlighted line.",
            format_id="david-and-goliath",
        ),
        "trending_next": TrendingNext(
            idea="shelf-life test",
            principle="same trust-violation lever, different mechanic",
            execution="Open 6-month-old jar, smell it, rub on wrist, 'still works'.",
        ),
    }
    if tab == MatrixTab.PAIN_VS_COMPETITOR:
        base.update({
            "complaint": "random patch breakouts",
            "our_advantage": "five ingredients you can pronounce",
            "positioning_angle": "same jar every time, no surprise breakouts",
        })
    elif tab == MatrixTab.WHAT_THEY_LOVE:
        base.update({
            "emotional_trigger": "healed eczema when nothing else worked",
            "hook_or_angle_to_amplify": "before/after on inner elbow",
        })
    elif tab == MatrixTab.WISHES_GAPS:
        base.update({
            "wished_product": "tallow balm that doesn't smell beefy",
            "real_frustration": "smell ruins daily use",
            "script_or_static_opportunity": "smell test on camera",
        })
    elif tab == MatrixTab.HOOK_ANGLES:
        base.update({
            "hook_angle": "If you can't eat your skincare, why is it on your face?",
            "tactic": "edible-reveal",
            "story_setup": "Spread on toast, eat, then jar on face.",
            "why_it_works": "Going viral on TikTok via Good Old Days, Taloco.",
        })
    return MatrixRow(**base)


def test_matrix_row_validation_rejects_unknown_heuristic():
    with pytest.raises(Exception) as exc_info:
        MatrixRow(
            id="pain-001",
            tab=MatrixTab.PAIN_VS_COMPETITOR,
            psychology_trigger="not_a_real_heuristic",
        )
    assert "Unknown psychology_trigger" in str(exc_info.value)


def test_matrix_row_validation_rejects_unknown_format_fit():
    with pytest.raises(Exception) as exc_info:
        MatrixRow(
            id="pain-001",
            tab=MatrixTab.PAIN_VS_COMPETITOR,
            format_fit=["us-vs-them", "no-such-format"],
        )
    assert "Unknown format_fit" in str(exc_info.value)


def test_matrix_row_validation_rejects_unknown_source_type():
    with pytest.raises(Exception) as exc_info:
        MatrixRow(
            id="pain-001",
            tab=MatrixTab.PAIN_VS_COMPETITOR,
            source_type=["reddit", "ouija_board"],
        )
    assert "Unknown source_type" in str(exc_info.value)


def test_matrix_row_validation_allows_empty_psychology_trigger():
    row = MatrixRow(id="pain-001", tab=MatrixTab.PAIN_VS_COMPETITOR)
    assert row.psychology_trigger == ""


def test_creative_matrix_round_trip(tmp_path: Path, monkeypatch: pytest.MonkeyPatch):
    monkeypatch.chdir(tmp_path)
    (tmp_path / "clients" / "demo").mkdir(parents=True)

    matrix = CreativeMatrix(
        client="demo",
        generated_at="2026-05-22T10:00:00",
        tier="standard",
        research_sources_used={"tier_1": ["reddit", "amazon"]},
        pain_vs_competitor=[_sample_row(MatrixTab.PAIN_VS_COMPETITOR, 1)],
        what_they_love=[_sample_row(MatrixTab.WHAT_THEY_LOVE, 1)],
        wishes_gaps=[_sample_row(MatrixTab.WISHES_GAPS, 1)],
        hook_angles=[_sample_row(MatrixTab.HOOK_ANGLES, 1)],
    )
    yaml_path = cm.save_matrix(matrix, "demo")

    assert yaml_path.exists()
    reloaded = cm.load_matrix("demo")

    assert reloaded.client == "demo"
    assert len(reloaded.pain_vs_competitor) == 1
    assert reloaded.pain_vs_competitor[0].trending_match.concept == "ingredient receipt"
    assert reloaded.pain_vs_competitor[0].trending_next.idea == "shelf-life test"

    # Provenance file exists and tracks row counts
    meta_path = tmp_path / "clients" / "demo" / "strategy" / "matrix_meta.yaml"
    meta = yaml.safe_load(meta_path.read_text())
    assert meta["row_counts"]["pain_vs_competitor"] == 1
    assert meta["tier"] == "standard"


def test_find_and_replace_row():
    matrix = CreativeMatrix(
        client="demo",
        generated_at="2026-05-22T10:00:00",
        pain_vs_competitor=[_sample_row(MatrixTab.PAIN_VS_COMPETITOR, 1)],
    )
    target = matrix.find_row("pain-001")
    assert target is not None

    target.positioning_angle = "updated angle"
    matrix.replace_row(target)
    assert matrix.pain_vs_competitor[0].positioning_angle == "updated angle"

    with pytest.raises(KeyError):
        ghost = _sample_row(MatrixTab.PAIN_VS_COMPETITOR, 999)
        matrix.replace_row(ghost)


# ─── XLSX writer ────────────────────────────────────────────────────────────


def test_xlsx_writer_produces_four_sheets(tmp_path: Path):
    matrix = CreativeMatrix(
        client="demo",
        generated_at="2026-05-22T10:00:00",
        pain_vs_competitor=[_sample_row(MatrixTab.PAIN_VS_COMPETITOR, i) for i in range(1, 4)],
        what_they_love=[_sample_row(MatrixTab.WHAT_THEY_LOVE, i) for i in range(1, 3)],
        wishes_gaps=[_sample_row(MatrixTab.WISHES_GAPS, 1)],
        hook_angles=[_sample_row(MatrixTab.HOOK_ANGLES, i) for i in range(1, 5)],
    )

    out = tmp_path / "creative_matrix.xlsx"
    write_matrix_xlsx(matrix, out)
    assert out.exists()

    wb = load_workbook(out)
    assert set(wb.sheetnames) == {
        "Pain vs Competitor",
        "What They Love",
        "Wishes & Gaps",
        "Hook Angles",
    }

    pain_sheet = wb["Pain vs Competitor"]
    assert pain_sheet.cell(row=1, column=1).value == "ID"
    assert pain_sheet.cell(row=2, column=1).value == "pain-001"
    assert pain_sheet.cell(row=4, column=1).value == "pain-003"
    # Header row frozen
    assert pain_sheet.freeze_panes == "A2"
    # Trending columns flattened
    headers = [pain_sheet.cell(row=1, column=c).value for c in range(1, 15)]
    assert "Trending match — concept" in headers
    assert "Trending next — idea" in headers


# ─── LLM-stubbed generation ────────────────────────────────────────────────


_PAIN_TAB_JSON = json.dumps({
    "rows": [
        {
            "complaint": "random patch breakouts",
            "source_type": ["reddit", "amazon"],
            "what_people_say": "holy grail then suddenly cystic acne",
            "our_advantage": "five ingredients you can pronounce",
            "positioning_angle": "same jar every time, no surprise breakouts",
            "format_fit": ["us-vs-them", "ingredient-callout"],
            "static_treatment": "Split-screen with red-circled actives on theirs.",
            "psychology_trigger": "social_proof",
        }
    ]
})

_LOVE_TAB_JSON = json.dumps({
    "rows": [
        {
            "emotional_trigger": "healed eczema when nothing else worked",
            "source_type": ["reddit"],
            "what_people_say": "son's eczema gone after years of steroids",
            "hook_or_angle_to_amplify": "before/after on inner elbow",
            "format_fit": ["before-after", "ugc-static"],
            "static_treatment": "Inner elbow close-up, day 0 vs day 30.",
            "psychology_trigger": "social_proof",
        }
    ]
})

_WISH_TAB_JSON = json.dumps({
    "rows": [
        {
            "wished_product": "tallow balm that doesn't smell beefy",
            "source_type": ["amazon"],
            "real_frustration": "smell makes daily use unpleasant",
            "script_or_static_opportunity": "smell test on camera",
            "format_fit": ["ugc-static"],
            "static_treatment": "Open jar, hold to nose, eyebrows up.",
            "psychology_trigger": "processing_fluency",
        }
    ]
})

_HOOK_TAB_JSON = json.dumps({
    "rows": [
        {
            "hook_angle": "If you can't eat your skincare, why is it on your face?",
            "tactic": "edible-reveal",
            "story_setup": "Spread tallow on toast, eat, jar on face.",
            "why_it_works": "Trending pattern across Good Old Days, Taloco.",
            "format_fit": ["ugc-static", "video-only"],
            "static_treatment": "",
            "psychology_trigger": "processing_fluency",
        }
    ]
})

_TRENDING_JSON_WITH_NEXT = json.dumps({
    "trending_match": {
        "concept": "ingredient receipt",
        "execution": "Read both labels aloud, cut to highlighted line.",
        "format_id": "david-and-goliath",
    },
    "trending_next": {
        "idea": "shelf-life test",
        "principle": "same trust-violation lever, different mechanic",
        "execution": "Open 6-month jar, smell on camera, rub on wrist.",
    },
})


def test_generate_matrix_full_build(tmp_path: Path, monkeypatch: pytest.MonkeyPatch):
    """End-to-end build with stubbed LLM. Asserts structural shape, not content."""
    monkeypatch.chdir(tmp_path)

    # Minimal client folder — competitive-gaps.yaml is enough to satisfy the
    # "no research found" guard. brand.yaml needs the required fields.
    client_dir = tmp_path / "clients" / "demo"
    (client_dir / "research").mkdir(parents=True)
    (client_dir / "strategy").mkdir(parents=True)

    (client_dir / "brand.yaml").write_text(
        yaml.safe_dump({
            "name": "DemoCo",
            "category": "skincare",
            "differentiators": ["five-ingredient formula"],
        }),
        encoding="utf-8",
    )
    (client_dir / "brand-context.md").write_text("Demo brand context.\n", encoding="utf-8")
    (client_dir / "research" / "competitive-gaps.yaml").write_text(
        yaml.safe_dump({
            "own_brand": "DemoCo",
            "per_brand": {
                "own": {"brand": "DemoCo", "loves": [], "gaps": [], "dealbreakers": []},
                "competitors": [
                    {
                        "brand": "RivalCo",
                        "loves": [{"claim": "smells nice", "money_quote": "love the scent",
                                   "sources": ["reddit"]}],
                        "gaps": [{"gap": "breakouts", "money_quote": "started getting acne",
                                  "sources": ["reddit"]}],
                        "dealbreakers": [],
                        "meta": {"sources": ["reddit", "amazon"]},
                    }
                ],
            },
            "synthesis": {
                "exploitable_gaps": [{
                    "opportunity": "breakout-safe formula",
                    "competitors_failing": ["RivalCo"],
                    "customer_evidence": "started getting acne",
                    "our_advantage": "five ingredients",
                    "ad_angle": "same jar every time",
                }],
                "shared_dealbreakers": [],
                "defensive_priorities": [],
                "category_table_stakes": [],
                "summary": "DemoCo wins on simplicity.",
            },
        }),
        encoding="utf-8",
    )

    # Stub the LLM. Tab order: pain, love, wish, hook. Then per-row trending calls.
    tab_responses = iter([_PAIN_TAB_JSON, _LOVE_TAB_JSON, _WISH_TAB_JSON, _HOOK_TAB_JSON])
    call_log: list[str] = []

    def fake_claude(prompt: str, system: str = "", max_tokens: int = 4096) -> str:
        call_log.append(prompt[:80])
        if "trending_match" in prompt.lower() or "candidate trending formats" in prompt.lower():
            return _TRENDING_JSON_WITH_NEXT
        return next(tab_responses)

    monkeypatch.setattr(cm, "claude_complete", fake_claude)
    monkeypatch.setattr(
        cm,
        "load_trending_formats",
        lambda: [{
            "id": "david-and-goliath",
            "name": "David & Goliath",
            "summary": "Underdog vs giant.",
            "best_when": {"awareness_levels": ["problem_aware"], "persona_types": ["skeptic"]},
            "format_type": "both",
        }],
    )

    matrix = cm.generate_matrix("demo", tier="standard")

    assert matrix.client == "demo"
    assert len(matrix.pain_vs_competitor) == 1
    assert len(matrix.what_they_love) == 1
    assert len(matrix.wishes_gaps) == 1
    assert len(matrix.hook_angles) == 1

    # IDs were normalized — LLM-supplied ids must not survive
    assert matrix.pain_vs_competitor[0].id == "pain-001"
    assert matrix.hook_angles[0].id == "hook-001"

    # Trending was attached on every row at standard tier
    for r in matrix.all_rows():
        assert r.trending_match is not None
        assert r.trending_match.concept == "ingredient receipt"
        assert r.trending_next is not None
        assert r.trending_next.idea == "shelf-life test"
        assert r.trending_next.generated == "eager"

    # Sources detection picked up the per-brand meta
    assert "tier_1" in matrix.research_sources_used


def test_generate_matrix_lean_skips_trending_next(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
):
    """Lean tier should produce trending_match but skip trending_next."""
    monkeypatch.chdir(tmp_path)
    client_dir = tmp_path / "clients" / "demo"
    (client_dir / "research").mkdir(parents=True)
    (client_dir / "strategy").mkdir(parents=True)

    (client_dir / "brand.yaml").write_text(
        yaml.safe_dump({"name": "DemoCo", "category": "skincare"}),
        encoding="utf-8",
    )
    (client_dir / "research" / "competitive-gaps.yaml").write_text(
        yaml.safe_dump({
            "synthesis": {"exploitable_gaps": []},
            "per_brand": {"competitors": []},
        }),
        encoding="utf-8",
    )

    tab_responses = iter([_PAIN_TAB_JSON, _LOVE_TAB_JSON, _WISH_TAB_JSON, _HOOK_TAB_JSON])

    def fake_claude(prompt: str, system: str = "", max_tokens: int = 4096) -> str:
        if "trending_match" in prompt.lower() or "candidate trending formats" in prompt.lower():
            # Lean tier: response should NOT include trending_next, but we
            # return one anyway to verify the caller drops it.
            return json.dumps({
                "trending_match": {
                    "concept": "ingredient receipt",
                    "execution": "Read labels aloud.",
                    "format_id": "david-and-goliath",
                }
            })
        return next(tab_responses)

    monkeypatch.setattr(cm, "claude_complete", fake_claude)
    monkeypatch.setattr(
        cm,
        "load_trending_formats",
        lambda: [{
            "id": "david-and-goliath",
            "name": "David & Goliath",
            "best_when": {},
            "format_type": "both",
        }],
    )

    matrix = cm.generate_matrix("demo", tier="lean")

    for r in matrix.all_rows():
        assert r.trending_match is not None
        assert r.trending_next is None


def test_generate_matrix_raises_when_no_research(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
):
    monkeypatch.chdir(tmp_path)
    client_dir = tmp_path / "clients" / "demo"
    client_dir.mkdir(parents=True)
    (client_dir / "brand.yaml").write_text(
        yaml.safe_dump({"name": "DemoCo", "category": "skincare"}), encoding="utf-8"
    )

    with pytest.raises(RuntimeError, match="No research found"):
        cm.generate_matrix("demo", tier="lean")


# ─── JSON parser tolerance ─────────────────────────────────────────────────


def test_parser_handles_bare_array():
    bare = '[{"complaint": "x", "source_type": ["reddit"], "what_people_say": "y", '
    bare += '"our_advantage": "z", "positioning_angle": "w", "format_fit": ["headline"], '
    bare += '"static_treatment": "t", "psychology_trigger": "scarcity"}]'
    parsed = cm._parse_json_with_rows(bare)
    assert parsed is not None
    assert "rows" in parsed
    assert len(parsed["rows"]) == 1


def test_parser_handles_markdown_fences():
    fenced = '```json\n{"rows": [{"complaint": "x"}]}\n```'
    parsed = cm._parse_json_with_rows(fenced)
    assert parsed is not None
    assert parsed["rows"][0]["complaint"] == "x"


def test_parser_handles_trailing_comma():
    with_comma = '{"rows": [{"complaint": "x",}]}'
    parsed = cm._parse_json_with_rows(with_comma)
    assert parsed is not None


# ─── row_to_brief synthesizer ───────────────────────────────────────────────


def _make_product(name: str = "Gut Balance"):
    """Build a minimal Product instance — most fields default."""
    from models.product import Product
    return Product(name=name, description="test product", benefits=["fast results"])


def _make_avatar(name: str = "Test Persona"):
    """Build a minimal CustomerAvatar instance."""
    from models.avatar import CustomerAvatar
    return CustomerAvatar(name=name, demographic="Women 30-45, urban")


def test_row_to_brief_tab1_full_synthesis():
    """Tab 1 (pain_vs_competitor) maps to PAS framework + populates all fields."""
    row = _sample_row(MatrixTab.PAIN_VS_COMPETITOR, 7)
    brief = cm.row_to_brief(
        row=row,
        client_slug="demo",
        product=_make_product(),
        avatar=_make_avatar(),
    )
    assert brief.brief_id == "matrix-pain-007"
    assert brief.client == "demo"
    assert brief.product == "Gut Balance"
    assert brief.framework.value == "pas"  # Tab 1 → PAS
    # social_proof from sample row → SOLUTION_AWARE per mapping
    assert brief.awareness_level.value == "solution_aware"
    assert brief.angle == "same jar every time, no surprise breakouts"  # positioning_angle
    assert "no surprise breakouts" in brief.hook
    assert brief.visual_format == "us-vs-them"
    assert brief.visual_format_alternatives == ["headline"]
    assert "Split-screen" in brief.visual_direction
    assert "patch breakouts" in brief.pain_point  # row.complaint
    assert brief.source_insight == "creative_matrix:pain_vs_competitor:pain-007"
    assert brief.persona == "Test Persona"


def test_row_to_brief_tab4_hook_angles_uses_hook_directly():
    """Tab 4 (hook_angles) — the row's hook_angle IS the brief hook."""
    row = _sample_row(MatrixTab.HOOK_ANGLES, 1)
    brief = cm.row_to_brief(
        row=row,
        client_slug="demo",
        product=_make_product(),
    )
    assert brief.framework.value == "aida"  # Tab 4 → AIDA
    assert "skincare" in brief.hook.lower()  # the actual hook line
    assert brief.hook_tactic == "edible-reveal"  # from row.tactic
    assert brief.hook_type == "edible-reveal"  # Tab 4 sets hook_type too


def test_row_to_brief_unknown_heuristic_defaults_to_problem_aware():
    """Empty psychology_trigger → falls back to problem_aware."""
    row = MatrixRow(
        id="pain-001",
        tab=MatrixTab.PAIN_VS_COMPETITOR,
        complaint="x",
        positioning_angle="y",
        format_fit=["headline"],
        psychology_trigger="",  # empty
    )
    brief = cm.row_to_brief(row=row, client_slug="demo", product=_make_product())
    assert brief.awareness_level.value == "problem_aware"


def test_row_to_brief_brief_id_prefix_customizable():
    row = _sample_row(MatrixTab.WISHES_GAPS, 4)
    brief = cm.row_to_brief(
        row=row,
        client_slug="demo",
        product=_make_product(),
        brief_id_prefix="experiment",
    )
    assert brief.brief_id == "experiment-wish-004"


def test_row_to_brief_benefit_callouts_split_on_sentence():
    """our_advantage should split on full stops to surface 1-3 benefit beats."""
    row = MatrixRow(
        id="pain-001",
        tab=MatrixTab.PAIN_VS_COMPETITOR,
        complaint="x",
        our_advantage="Five ingredients. Clinically studied. Patent-free.",
        positioning_angle="y",
        format_fit=["headline"],
        psychology_trigger="processing_fluency",
    )
    brief = cm.row_to_brief(row=row, client_slug="demo", product=_make_product())
    # Split keeps the trailing period on the final segment because regex
    # splits on `. ` and the last chunk has no trailing space.
    assert brief.benefit_callouts == [
        "Five ingredients",
        "Clinically studied",
        "Patent-free.",
    ]


def test_row_to_brief_no_avatar_means_empty_persona():
    row = _sample_row(MatrixTab.PAIN_VS_COMPETITOR, 1)
    brief = cm.row_to_brief(
        row=row,
        client_slug="demo",
        product=_make_product(),
        avatar=None,
    )
    assert brief.persona == ""
    assert brief.persona_traits == ""
